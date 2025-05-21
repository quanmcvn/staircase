import os
import signal
import sys
import time
from datetime import datetime

from openpyxl import Workbook

from src.encoding.car_sequencing import CarSequencing
from src.include.addline import write_full
from src.include.common import AuxVariable, AddClause, myrange_inclusive
from src.include.common import write_to_cell, pos_2d_to_pos_excel


def handler(signum, frame):
	raise TimeoutError("Function execution time exceeded the limit")


def read_car_sequencing_input_from_file(filename: str) -> tuple[
	int, int, int, list[tuple[int, int]], list[CarSequencing.ClassDesc]]:
	with open(filename, 'r') as file:
		num_cars, num_options, num_classes = list(map(int, file.readline().split()))
		caps = list(map(int, file.readline().split()))
		window_sizes = list(map(int, file.readline().split()))
		option_cap: list[tuple[int, int]] = [(caps[_], window_sizes[_]) for _ in range(len(caps))]
		class_desc: list[CarSequencing.ClassDesc] = []
		for _ in range(num_classes):
			list_int = list(map(int, file.readline().split()))
			index = list_int[0]
			demand = list_int[1]
			is_option_required: list[bool] = [True if list_int[__] == 1 else False for __ in range(2, len(list_int))]
			class_desc.append(CarSequencing.ClassDesc(index, demand, is_option_required))
	return num_cars, num_options, num_classes, option_cap, class_desc


def eval_window(x: list[int], window_size: int, cap: int):
	now = 0
	for i in range(0, len(x)):
		if i >= window_size:
			if now > cap:
				return False
			now -= 1 if x[i - window_size] != 0 else 0
		now += 1 if x[i] != 0 else 0
	if now > cap:
		return False
	return True


def parse_results(filename: str, num_cars: int, class_desc: list[CarSequencing.ClassDesc]) -> tuple[
	list[int], list[list[int]]]:
	m = len(class_desc)
	with (open(filename, 'r') as file):
		list_int: list[int] = []
		for line in file.readlines():
			for seq in line.split():
				try:
					list_int.append(int(seq))
				except ValueError:
					pass
		car_list = []
		for i in range(num_cars):
			car_list.append(list_int[i * m:i * m + m])
		chosen_option_list = []
		chosen_list = []
		for i in range(num_cars):
			temp = [j for j in car_list[i] if j > 0]
			if len(temp) != 1:
				raise RuntimeError(f"NOT satisfied 1 class per car! {temp}")
			positive = (temp[0] - 1) % m
			chosen_list.append(positive)
			this_class_desc = [1 if j else 0 for j in class_desc[positive].is_option_required]
			chosen_option_list.append(this_class_desc)
	return chosen_list, chosen_option_list


def run_car_sequencing(type_: str, filename: str) -> tuple[float | None, str, int, int]:
	signal.signal(signal.SIGALRM, handler)
	timeout = 30 * 60
	signal.alarm(timeout)
	num_cars_, num_options_, num_classes_, option_cap_, class_desc_ = read_car_sequencing_input_from_file(
		filename)
	# print(num_cars_, num_options_, num_classes_, option_cap_, sep='\n')
	# for class_desc__ in class_desc_:
	# 	print(class_desc__)
	num_vars = -1
	add_clause = AddClause([])
	using_kissat = True
	try:
		start_time = time.perf_counter()
		cs = CarSequencing()
		# cars = [i for i in myrange_inclusive(1, num_cars_)]
		aux = AuxVariable(1)
		if type_ == 'staircase':
			func = cs.encode_car_sequencing_staircase
		elif type_ == 'binomial':
			func = cs.encode_car_sequencing_binomial
		elif type_ == 'staircase_binomial':
			func = cs.encode_car_sequencing_staircase_binomial
		elif type_ == 'nsc':
			func = cs.encode_car_sequencing_nsc
		elif type_ == 'cpaior_2014':
			func = cs.encode_car_sequencing_cpaior_2014
		elif type_ == 'cpaior_2014_binomial':
			func = cs.encode_car_sequencing_cpaior_2014_binomial
		elif type_ == 'cpaior_2014_c_s':
			func = cs.encode_car_sequencing_cpaior_2014_c_s
		else:
			raise RuntimeError(f"unknown type {type_}")
		func(num_cars_, num_classes_, num_options_, option_cap_, class_desc_, aux, add_clause)
		# for clause in add_clause.get_clause():
		# 	print(clause)
		solver_return = 'unknown'
		num_vars = aux.get_last_used_var()
		filename = os.path.splitext(os.path.basename(filename))[0]
		cnf_file = f'/tmp/amk/{type_}_{filename}'
		solver_output = f'{cnf_file}_result'
		# print(f"took {elapsed_time_ms:.2f} (ms)")
		# os.system(f"cat {cnf_file}")
		# print(f"kissat -q {cnf_file} > {solver_output}")
		print(type_)
		print(f"total {num_vars} {add_clause.get_added_clause()}")
		write_full(num_vars, add_clause.get_clause(), cnf_file)
		if using_kissat:
			ret = os.system(f"kissat -q --time={timeout} {cnf_file} > {solver_output}")
			print(f"kissat ret: {ret}")
		else:
			ret = os.system(f"glucose-syrup -maxnbthreads=6 -cpu-lim={timeout * 6} {cnf_file} > {solver_output}")
			print(f"glucose-syrup ret: {ret}")

		end_time = time.perf_counter()

		elapsed_time_ms = (end_time - start_time) * 1000
		ok_time = True
		if using_kissat:
			if ret == 2560:  # SAT
				solver_return = 'SAT'
				chosen_class, chosen_class_option = parse_results(solver_output, num_cars_, class_desc_)
				# for i in range(num_cars_):
				# 	print(chosen_class[i], chosen_class_option[i])
				ok_option_cap = True
				ok_demand = True
				for o in range(0, len(option_cap_)):
					cap, window_size = option_cap_[o]
					option_list = []
					for i in range(num_cars_):
						option_list.append(chosen_class_option[i][o])
					if not eval_window(option_list, window_size, cap):
						# raise RuntimeError(f"Chosen list NOT satisfy option cap!!!")
						ok_option_cap = False
				demand_count = [0] * len(class_desc_)
				for i in range(num_cars_):
					demand_count[chosen_class[i]] += 1
				for i in range(len(class_desc_)):
					if demand_count[i] != class_desc_[i].demand:
						# raise RuntimeError(f"Chosen list NOT satisfy demand!!!")
						ok_demand = False
				if ok_option_cap and ok_demand:
					print("Chosen list satisfied.")
					for i in range(len(chosen_class)):
						print(chosen_class[i], end=' ')
						for j in chosen_class_option[i]:
							print(j, end=' ')
						print()
				if not ok_option_cap:
					print(f"Chosen list NOT satisfy option cap!!!")
				if not ok_demand:
					print(f"Chosen list NOT satisfy demand!!!")
			elif ret == 5120:  # UNSAT
				print("UNSAT")
				solver_return = 'UNSAT'
			else:
				if ret == 0: # timeout
					ok_time = False
		else:
			if ret == 256:
				solver_return = 'unknown'
				ok_time = False
			else:
				pass
		# return elapsed_time_ms, num_vars, add_clause.get_added_clause()

		print(f"took {elapsed_time_ms:.2f} (ms)")
		print('--------------------------------')
		if ok_time:
			return elapsed_time_ms, solver_return, num_vars, add_clause.get_added_clause()
		else:
			return None, 'timeout', num_vars, add_clause.get_added_clause()
	except TimeoutError as te:
		print(te)
		return None, 'timeout', num_vars, add_clause.get_added_clause()
	except RecursionError as re:
		print(re)
		return None, 'timeout', num_vars, add_clause.get_added_clause()
	except OSError as e:
		print(e)
		return None, 'timeout', num_vars, add_clause.get_added_clause()
	finally:
		signal.alarm(0)


class DataToXlsx:
	total_variable = 'total_variable'
	clause = 'clause'
	time = 'time (ms)'
	sat_status = 'sat/unsat/timeout'

	def __init__(self, excel_file_name: str, name_list: list[str]):
		if not os.path.exists("out/excel"):
			os.makedirs("out/excel")
		self.excel_file_name = "out/excel/" + excel_file_name
		if os.path.exists(self.excel_file_name):
			print(f"file {self.excel_file_name} exist! aborting...")
			sys.exit(2)
		self.book = Workbook()
		for sheet_name in self.book.sheetnames:
			self.book.remove(self.book[sheet_name])
		self.book.create_sheet('Results')
		self.sheet = self.book['Results']
		self.name_to_column: dict[str, int] = {
			'filename': 1,
		}
		self.row_dict: dict[str, int] = {
			'filename': 3,
		}
		self.offset_dict: dict[str, int] = {
			'filename': 0,
			DataToXlsx.total_variable: 0,
			DataToXlsx.clause: 1,
			DataToXlsx.time: 2,
			DataToXlsx.sat_status: 3,
		}
		write_to_cell(self.sheet.cell(2, 1), 'filename')
		# self.sheet.cell(2, 1).value = 'n'
		# self.sheet.cell(2, 1).alignment = Alignment(horizontal='center')
		# self.sheet.cell(2, 2).value = 'k'

		for i in range(len(name_list)):
			y: int = 2 + i * 4
			self.name_to_column[name_list[i]] = y
			self.sheet.merge_cells(f"{pos_2d_to_pos_excel(1, y)}:{pos_2d_to_pos_excel(1, y + 2)}")
			write_to_cell(self.sheet.cell(1, y), name_list[i])
			# self.sheet.cell(1, y).alignment = Alignment(horizontal='center')
			# self.sheet.cell(1, y).value = name_list[i]
			for name_offset, offset in self.offset_dict.items():
				write_to_cell(self.sheet.cell(2, y + offset), name_offset)
			# self.sheet.cell(2, y + offset).value = name_offset
			self.row_dict[name_list[i]] = 3
		self.book.save(self.excel_file_name)

	def get_column(self, name: str) -> int:
		if not (name in self.name_to_column):
			raise RuntimeError(f"{name} not in name_to_column")
		return self.name_to_column.get(name)

	def get_row(self, name: str) -> int:
		ret: int = self.row_dict.get(name)
		self.row_dict[name] += 1
		return ret

	def write_more(self, name: str, result: dict[str, str | int | float]) -> None:
		row: int = self.get_row(name)
		col: int = self.get_column(name)
		for name_offset, value in result.items():
			offset: int = self.offset_dict[name_offset]
			write_to_cell(self.sheet.cell(row, col + offset), value)
		# current_cell = self.sheet.cell(row, col + offset)
		# current_cell.value = value
		# current_cell.alignment = Alignment(horizontal='center')
		self.book.save(self.excel_file_name)


def get_filenames_from_directory(directory: str) -> list[str]:
	filenames = []
	for filename in os.listdir(directory):
		file_path = os.path.join(directory, filename)
		if os.path.isfile(file_path):  # Check if it is a file
			filenames.append(file_path)
	return filenames


def main():
	filenames = []
	# for i in myrange_inclusive(1, 9):
	# 	filenames.append(f'./inputs/car_sequencing/hard_prob/pb_100_{i:02d}.txt')
	# filenames.append(f'./inputs/car_sequencing/hard_prob/pb_200_03.txt')
	# filenames.append(f'./inputs/car_sequencing/hard_prob/pb_200_05.txt')
	# filenames.append(f'./inputs/car_sequencing/hard_prob/pb_400_03.txt')
	# for x in get_filenames_from_directory('./inputs/car_sequencing/hard_prob'):
	# 	filenames.append(x)
	# for i in myrange_inclusive(61, 61):
	# 	filenames.append(f'./inputs/car_sequencing/prob-200-{i:02d}.txt')
	filenames.append('./inputs/car_sequencing/prob-sample.txt')
	# filenames = get_filenames_from_directory('./inputs/car_sequencing/')

	filenames.sort()
	# filenames.append('inputs/car_sequencing/prob-200-1.txt')
	# filenames.append('inputs/car_sequencing/prob-200-1_temp.txt')
	# filenames.append('inputs/car_sequencing/prob-200-2.txt')
	# filenames.append('inputs/car_sequencing/prob-200-3.txt')
	# filenames.append('inputs/car_sequencing/prob-200-4.txt')
	# filenames.append('inputs/car_sequencing/prob-200-5.txt')
	to_test = []
	for x in ["staircase_binomial"]:
		to_test.append(x)
	# for x in ["cpaior_2014_c_s"]:
	# 	to_test.append(x)
	# for x in ["binomial"]:
	# 	to_test.append(x)
	# for x in ["nsc"]:
	# 	to_test.append(x)
	# for x in ["cpaior_2014"]:
	# 	to_test.append(x)
	# for x in ["staircase", "cpaior_2014"]:
	# 	to_test.append(x)

	time_now = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
	excel_file_name = f"results_{time_now}.xlsx"
	writer: DataToXlsx = DataToXlsx(excel_file_name, to_test)

	for filename in filenames:
		print(f"running {filename}")
		writer.write_more('filename', {'filename': filename})
		for name in to_test:
			result_dict: dict[str, str | int | float] = {
				DataToXlsx.total_variable: "",
				DataToXlsx.clause: "",
				DataToXlsx.time: "",
			}
			elapsed_time_ms, solver_return, num_var, num_clause = run_car_sequencing(name, filename)
			if elapsed_time_ms is None:
				result_dict[DataToXlsx.time] = "timeout"
			else:
				result_dict[DataToXlsx.time] = f"{elapsed_time_ms:.3f}"
			if num_var == -1:
				result_dict[DataToXlsx.total_variable] = ""
				result_dict[DataToXlsx.clause] = ""
			else:
				result_dict[DataToXlsx.total_variable] = num_var
				result_dict[DataToXlsx.clause] = num_clause
			result_dict[DataToXlsx.sat_status] = solver_return

			writer.write_more(name, result_dict)


if __name__ == '__main__':
	main()
