import os
import signal
import sys
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from datetime import datetime
from typing import Any
from src.include import function_template


def handler(signum, frame):
	raise TimeoutError("Function execution time exceeded the limit")


def run(command: str) -> int:
	print(command)
	return os.system(command)


def test(name: str, n: int, k: int, m: int, time_limit: int) -> tuple[float | None, int, int]:
	signal.signal(signal.SIGALRM, handler)
	signal.alarm(time_limit)
	if "range" in name:
		cannon_name = f"{name}_{n}_{k}_{m}.cnf"
	else:
		cannon_name = f"{name}_{n}_{k}.cnf"
	if not os.path.exists("/tmp/amk"):
		os.makedirs("/tmp/amk")
	filename = f"/tmp/amk/{cannon_name}"
	print(f"{name} {n} {k}")
	num_var: int = -1
	num_clause: int = -1
	try:
		start_time = time.perf_counter()
		if "range" in name:
			func = function_template.get_basic_range_function(name)
			num_var, num_clause = func(n, k, m, filename)
		else:
			func = function_template.get_basic_function(name)
			num_var, num_clause = func(n, k, filename)
		os.system(f"head -n1 {filename}")
		if not os.path.exists("/tmp/amk/kissat_output"):
			os.makedirs("/tmp/amk/kissat_output")
		run(f"kissat -n --time={time_limit} {filename} > /tmp/amk/kissat_output/output_{cannon_name}.txt")
		end_time = time.perf_counter()
		elapsed_time_ms = (end_time - start_time) * 1000
		print(f"took {elapsed_time_ms:.2f} (ms)")
	except TimeoutError as te:
		print(te)
		return None, num_var, num_clause
	except RecursionError as re:
		print(re)
		return None, num_var, num_clause
	except OSError as e:
		print(e)
		return None, num_var, num_clause
	finally:
		signal.alarm(0)
	# run(f"rm /tmp/amk/{cannon_name}")
	print("-" * 100)
	return elapsed_time_ms, num_var, num_clause


# def get_all_string_in_file(file: str) -> list[list[str]]:
# 	args = []
# 	file = open(f'{file}', 'r')
# 	for line in file:
# 		arg = [tok for tok in line.split()]
# 		args.append(arg)
# 	file.close()
# 	return args


def get_all_number_in_file(file: str) -> list[list[int]]:
	args = []
	file = open(f'{file}', 'r')
	for line in file:
		arg = [int(num) for num in line.split()]
		args.append(arg)
	file.close()
	return args


def number_to_column_letter(n: int) -> str:
	result = ""
	while n > 0:
		n, remainder = divmod(n - 1, 26)
		result = chr(65 + remainder) + result

	return result


def pos_2d_to_pos_excel(x: int, y: int) -> str:
	ret = number_to_column_letter(y) + str(x)
	# print(f"({x}, {y}) -> {ret}")
	return ret


def write_to_cell(cell, value: Any):
	cell.value = value
	cell.alignment = Alignment(horizontal="center")


class DataToXlsx:
	total_variable = 'total_variable'
	new_variable = 'new_variable'
	clause = 'clause'
	time = 'time'

	def __init__(self, excel_file_name: str, name_list: list[str], is_range: bool):
		if not os.path.exists("out/excel"):
			os.makedirs("out/excel")
		self.excel_file_name = "out/excel/" + excel_file_name
		if os.path.exists(self.excel_file_name):
			print(f"file {self.excel_file_name} exist! aborting...")
			sys.exit(2)
		self.book = Workbook()
		for sheet_name in self.book.sheetnames:
			self.book.remove(self.book[sheet_name])
		self.book.create_sheet('Results')
		self.sheet = self.book['Results']
		self.name_to_column: dict[str, int] = {
			'n': 1,
			'k': 2,
		}
		if is_range:
			self.name_to_column['m'] = 3
		self.row_dict: dict[str, int] = {
			'n': 3,
			'k': 3,
			'm': 3
		}
		self.offset_dict: dict[str, int] = {
			'n': 0,
			'k': 0,
			'm': 0,
			DataToXlsx.total_variable: 0,
			DataToXlsx.new_variable: 1,
			DataToXlsx.clause: 2,
			DataToXlsx.time: 3,
		}
		write_to_cell(self.sheet.cell(2, 1), 'n')
		write_to_cell(self.sheet.cell(2, 2), 'k')
		if is_range:
			write_to_cell(self.sheet.cell(2, 3), 'm')
		# self.sheet.cell(2, 1).value = 'n'
		# self.sheet.cell(2, 1).alignment = Alignment(horizontal='center')
		# self.sheet.cell(2, 2).value = 'k'

		for i in range(len(name_list)):
			y: int = 3 + i * 4
			if is_range:
				y = y + 1
			self.name_to_column[name_list[i]] = y
			self.sheet.merge_cells(f"{pos_2d_to_pos_excel(1, y)}:{pos_2d_to_pos_excel(1, y + 3)}")
			write_to_cell(self.sheet.cell(1, y), name_list[i])
			# self.sheet.cell(1, y).alignment = Alignment(horizontal='center')
			# self.sheet.cell(1, y).value = name_list[i]
			for name_offset, offset in self.offset_dict.items():
				write_to_cell(self.sheet.cell(2, y + offset), name_offset)
			# self.sheet.cell(2, y + offset).value = name_offset
			self.row_dict[name_list[i]] = 3
		self.book.save(self.excel_file_name)

	def get_column(self, name: str) -> int:
		if not (name in self.name_to_column):
			raise RuntimeError(f"{name} not in name_to_column")
		return self.name_to_column.get(name)

	def get_row(self, name: str) -> int:
		ret: int = self.row_dict.get(name)
		self.row_dict[name] += 1
		return ret

	def write_more(self, name: str, result: dict[str, str | int | float]) -> None:
		row: int = self.get_row(name)
		col: int = self.get_column(name)
		for name_offset, value in result.items():
			offset: int = self.offset_dict[name_offset]
			write_to_cell(self.sheet.cell(row, col + offset), value)
		# current_cell = self.sheet.cell(row, col + offset)
		# current_cell.value = value
		# current_cell.alignment = Alignment(horizontal='center')
		self.book.save(self.excel_file_name)


# def write_to_excel(excel_file_path: str, result_dict: dict[str, str | int | float]):
# 	excel_results = [result_dict]
#
# 	# Write the results to an Excel file
# 	if not os.path.exists("out/excel"):
# 		os.makedirs("out/excel")
#
# 	df: pd.DataFrame = pd.DataFrame(excel_results)
#
# 	# Check if the file already exists
# 	if os.path.exists(excel_file_path):
# 		try:
# 			book = load_workbook(excel_file_path)
# 		except BadZipFile:
# 			book = Workbook()  # Create a new workbook if the file is not a valid Excel file
# 		# Check if the 'Results' sheet exists
# 		if 'Results' not in book.sheetnames:
# 			book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist
#
# 		sheet = book['Results']
#
# 		for row in dataframe_to_rows(df, index=False, header=False):
# 			sheet.append(row)
#
# 		book.save(excel_file_path)
#
# 	else:
# 		df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)


def main(run_type: str):
	to_test: list[str] = []
	# for x in ["nsc_hybrid", "nsc_at_least_k", "sc_at_least_k", "binary_at_least_k", "commander_at_least_k", "product_at_least_k", "binomial_at_least_k", "pblib_bdd_at_least_k", "pblib_card_at_least_k"]:
	# 	to_test.append(x)
	# for x in ["nsc_at_most_k", "sc_at_most_k", "binary_at_most_k", "commander_at_most_k", "product_at_most_k", "binomial_at_most_k", "pblib_bdd_at_most_k", "pblib_card_at_most_k"]:
	# 	to_test.append(x)
	# for x in ["nsc_exactly_k", "sc_exactly_k", "binary_exactly_k", "commander_exactly_k", "product_exactly_k", "binomial_exactly_k", "pblib_bdd_exactly_k", "pblib_card_exactly_k"]:
	# 	to_test.append(x)

	# for x in ["nsc_hybrid", "nsc_at_least_k", "sc_at_least_k", "binary_at_least_k", "pblib_bdd_at_least_k", "pblib_card_at_least_k", "pblib_add_at_least_k"]:
	# 	to_test.append(x)
	# for x in ["nsc_at_most_k", "sc_at_most_k", "binary_at_most_k", "pblib_bdd_at_most_k", "pblib_card_at_most_k", "pblib_add_at_most_k"]:
	# 	to_test.append(x)
	# for x in ["nsc_exactly_k", "sc_exactly_k", "binary_exactly_k", "pblib_bdd_exactly_k", "pblib_card_exactly_k", "pblib_add_exactly_k"]:
	# 	to_test.append(x)

	# for x in ["pblib_bdd_at_most_k", "pblib_bdd_at_least_k", "pblib_bdd_range"]:
	# 	to_test.append(x)
	# for x in ["pblib_card_at_most_k", "pblib_card_at_least_k", "pblib_card_range"]:
	# 	to_test.append(x)
	for x in ["pblib_bdd_exactly_k"]:
		to_test.append(x)
	# for x in ["nsc_at_most_k", "sc_at_most_k", "binary_at_most_k", "pblib_bdd_at_most_k", "pblib_card_at_most_k", "pblib_add_at_most_k"]:
	# 	to_test.append(x)
	# for x in ["nsc_exactly_k", "sc_exactly_k", "binary_exactly_k", "pblib_bdd_exactly_k", "pblib_card_exactly_k", "pblib_add_exactly_k"]:
	# 	to_test.append(x)

	# writer: DataToXlsx = DataToXlsx("aaa.xlsx", ['nsc_alk', 'nsc_amk', 'binary_ek'])
	time_now = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
	nks = get_all_number_in_file("input.txt")
	is_range: bool = False
	for nk in nks:
		if len(nk) > 2:
			is_range = True
			break
	excel_file_name = f"results_{time_now}.xlsx"
	writer: DataToXlsx = DataToXlsx(excel_file_name, to_test, is_range)
	timeout = 600
	for nk in nks:
		n = nk[0]
		k = nk[1]
		m = -1
		is_range_local = False
		if len(nk) > 2:
			is_range_local = True
		writer.write_more('n', {'n': n})
		writer.write_more('k', {'k': k})
		if is_range_local:
			m = nk[2]
			writer.write_more('m', {'m': m})
		else:
			if is_range:
				writer.write_more('m', {'m': ""})
		for name in to_test:
			result_dict: dict[str, str | int | float] = {
				DataToXlsx.total_variable: "",
				DataToXlsx.new_variable: "",
				DataToXlsx.clause: "",
				DataToXlsx.time: "",
			}
			if is_range_local and "range" not in name:
				print(n, k, m, name, "range input but not range")
				writer.write_more(name, result_dict)
				continue
			if not is_range_local and "range" in name:
				print(n, k, m, name, "range but not range input")
				writer.write_more(name, result_dict)
				continue
			elapsed_time_ms, num_var, num_clause = test(name, n, k, m, timeout)
			if elapsed_time_ms is None:
				result_dict[DataToXlsx.time] = "timeout"
			else:
				result_dict[DataToXlsx.time] = f"{elapsed_time_ms:.3f}"
			if num_var == -1:
				result_dict[DataToXlsx.total_variable] = ""
				result_dict[DataToXlsx.new_variable] = ""
				result_dict[DataToXlsx.clause] = ""
			else:
				result_dict[DataToXlsx.total_variable] = num_var
				result_dict[DataToXlsx.new_variable] = num_var - n
				result_dict[DataToXlsx.clause] = num_clause

			writer.write_more(name, result_dict)
		print()
	pass


if __name__ == '__main__':
	main("")
