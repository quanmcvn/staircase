import os
import signal
import sys
import datetime
import gc
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from datetime import datetime
from typing import Any

from encoding.NRP.nurse_roostering_encoding import NurseRosteringEncoding, NurseRosteringConfig
from src.encoding.staircase_encoding import StaircaseEncoding
from src.include.addline import write_full
from src.include.common import myrange_inclusive, cl, AuxVariable, AddClause


def handler(signum, frame):
	raise TimeoutError("Function execution time exceeded the limit")


def run(command: str) -> int:
	print(command)
	return os.system(command)


def eval_window(x: str, value: str, window_size: int, floor, cap):
	if len(x) < window_size:
		return True
	now = 0
	if floor is None:
		floor = 0
	if cap is None:
		cap = window_size
	for i in range(0, len(x)):
		if i >= window_size:
			if not (floor <= now <= cap):
				return False
			now -= 1 if x[i - window_size] == value else 0
		now += 1 if x[i] == value else 0
	if not (floor <= now <= cap):
		return False
	return True


def eval_window_lower_bound(x: str, value: str, window_size: int, floor: int):
	return eval_window(x, value, window_size, floor, None)


def eval_window_upper_bound(x: str, value: str, window_size: int, cap: int):
	return eval_window(x, value, window_size, None, cap)


def run_nurse_rostering(name: str, nurse: int, day: int, time_limit: int) -> tuple[float | None, str, int, int]:
	signal.signal(signal.SIGALRM, handler)
	signal.alarm(time_limit)
	cannon_name = f"nurse_rostering_{name}"
	if not os.path.exists("tmp"):
		os.makedirs("tmp")
	cnf_file = f"tmp/{cannon_name}.cnf"
	print(f"{name} {nurse} {day}")
	aux = AuxVariable(1)
	clause = []
	add_clause = AddClause(clause)
	total_variable = -1
	total_clause = -1
	try:
		start_encoding_time = time.perf_counter()
		nr_config = NurseRosteringConfig(nurse, day, aux, add_clause, name)
		nr = NurseRosteringEncoding(nr_config)
		nr.encode()
		write_full(aux.get_total_added_var(), add_clause.get_clause(), cnf_file)
		os.system(f"head -n1 {cnf_file}")
		solver_output = f"tmp/output_{cannon_name}.txt"
  
		end_encoding_time = time.perf_counter()
		encoding_elapsed_time_ms = (end_encoding_time - start_encoding_time) * 1000
  
		start_solving_time = time.perf_counter()
		# For kissat solver
		ret = run(f"./kissat -q --time={time_limit} {cnf_file} > {solver_output}")
  
		# For kissat in 2024 competition
		# ret = run(f"./kissat_compe_2024 --time={time_limit} {cnf_file} > {solver_output}")

		# For Cadical solver
		# ret = run(f"./cadical -q -t {time_limit} {cnf_file} > {solver_output}")
  
		# For Glucose solver
		# ret = run(f"./glucose-syrup -cpu-lim={time_limit} -nthreads=8 -verb=0 {cnf_file} {solver_output}")
  
		end_solving_time = time.perf_counter()
		solving_elapsed_time_ms = (end_solving_time - start_solving_time) * 1000

		total_elapsed_time_ms = encoding_elapsed_time_ms + solving_elapsed_time_ms

		total_variable = aux.get_total_added_var()
		total_clause = add_clause.get_added_clause()
		del nr
		del clause
		gc.collect()

		solver_return = ''

		ok_time = True
		if ret == 2560:  # SAT
			solver_return = 'SAT'
			# test_result(solver_output, nurse, day)
		elif ret == 5120:  # UNSAT
			print("UNSAT")
			solver_return = 'UNSAT'
		else:
			if ret == 0:  # timeout
				ok_time = False

		print(f"Encoding elapsed time: {encoding_elapsed_time_ms:.2f} ms")
		print(f"Solving elapsed time: {solving_elapsed_time_ms:.2f} (ms)")
		print(f"Total elapsed time: {total_elapsed_time_ms:.2f} (ms)") 
		if ok_time:
			return total_elapsed_time_ms, solver_return, total_variable, total_clause
		else:
			return None, 'timeout', total_variable, total_clause
	except TimeoutError as te:
		print(te)
		return None, 'timeout', total_variable, total_clause
	except RecursionError as re:
		print(re)
		return None, 'timeout', total_variable, total_clause
	except OSError as e:
		print(e)
		return None, 'timeout', total_variable, total_clause
	finally:
		signal.alarm(0)


def get_all_number_in_file(file: str) -> list[list[int]]:
	args = []
	file = open(f'{file}', 'r')
	for line in file:
		arg = [int(num) for num in line.split()]
		args.append(arg)
	file.close()
	return args


def number_to_column_letter(n: int) -> str:
	result = ""
	while n > 0:
		n, remainder = divmod(n - 1, 26)
		result = chr(65 + remainder) + result

	return result


def pos_2d_to_pos_excel(x: int, y: int) -> str:
	ret = number_to_column_letter(y) + str(x)
	# print(f"({x}, {y}) -> {ret}")
	return ret


def write_to_cell(cell, value: Any):
	cell.value = value
	cell.alignment = Alignment(horizontal="center")


class DataToXlsx:
	total_variable = 'total_variable'
	clause = 'clause'
	time = 'time (ms)'
	sat_status = 'sat/unsat/timeout'

	def __init__(self, excel_file_name: str, name_list: list[str]):
		if not os.path.exists("out/excel"):
			os.makedirs("out/excel")
		self.excel_file_name = "out/excel/" + excel_file_name
		if os.path.exists(self.excel_file_name):
			print(f"file {self.excel_file_name} exist! aborting...")
			sys.exit(2)
		self.book = Workbook()
		for sheet_name in self.book.sheetnames:
			self.book.remove(self.book[sheet_name])
		self.book.create_sheet('Results')
		self.sheet = self.book['Results']
		self.name_to_column: dict[str, int] = {
			'nurse': 1,
			'day': 2,
		}
		self.row_dict: dict[str, int] = {
			'nurse': 3,
			'day': 3,
		}
		self.offset_dict: dict[str, int] = {
			'nurse': 0,
			'day': 0,
			DataToXlsx.total_variable: 0,
			DataToXlsx.clause: 1,
			DataToXlsx.time: 2,
			DataToXlsx.sat_status: 3,
		}
		write_to_cell(self.sheet.cell(2, 1), 'nurse')
		write_to_cell(self.sheet.cell(2, 2), 'day')

		for i in range(len(name_list)):
			y: int = len(self.row_dict) + 1 + i * 4
			self.name_to_column[name_list[i]] = y
			self.sheet.merge_cells(f"{pos_2d_to_pos_excel(1, y)}:{pos_2d_to_pos_excel(1, y + 3)}")
			write_to_cell(self.sheet.cell(1, y), name_list[i])
			for name_offset, offset in self.offset_dict.items():
				write_to_cell(self.sheet.cell(2, y + offset), name_offset)
			self.row_dict[name_list[i]] = 3
		self.book.save(self.excel_file_name)

	def get_column(self, name: str) -> int:
		if not (name in self.name_to_column):
			raise RuntimeError(f"{name} not in name_to_column")
		return self.name_to_column.get(name)

	def get_row(self, name: str) -> int:
		ret: int = self.row_dict.get(name)
		self.row_dict[name] += 1
		return ret

	def write_more(self, name: str, result: dict[str, str | int | float]) -> None:
		row: int = self.get_row(name)
		col: int = self.get_column(name)
		for name_offset, value in result.items():
			offset: int = self.offset_dict[name_offset]
			write_to_cell(self.sheet.cell(row, col + offset), value)
		self.book.save(self.excel_file_name)


def parse_result(filename: str, nurse: int, day: int):
	with (open(filename, 'r') as file):
		list_int: list[int] = []
		for line in file.readlines():
			for seq in line.split():
				try:
					list_int.append(int(seq))
				except ValueError:
					pass
	nurse_list = []
	list_int.reverse()
	for i in range(nurse):
		nurse_day_list = []
		for j in range(day):
			nurse_day_shift_list = []
			for k in range(4):
				nurse_day_shift_list.append(list_int.pop())
			nurse_day_list.append(nurse_day_shift_list)
		nurse_list.append(nurse_day_list)
	chosen_list = []
	num_to_shift = ['D', 'E', 'N', 'O']
	for i in range(nurse):
		chosen_nurse_list = []
		for j in range(day):
			temp = [k for k in nurse_list[i][j] if k > 0]
			if len(temp) != 1:
				raise RuntimeError(f"NOT satisfied 1 shift per day per nurse! {temp}")
			positive = (temp[0] - 1) % 4
			chosen_nurse_list.append(num_to_shift[positive])
		chosen_list.append(chosen_nurse_list)
	return chosen_list


def test_result(filename: str, nurse: int, day: int):
	chosen_list = parse_result(filename, nurse, day)
	# for i in range(num_cars_):
	# 	print(chosen_class[i], chosen_class_option[i])
	for nurse_id in range(nurse):
		nurse_shifts = ''
		for shift in chosen_list[nurse_id]:
			nurse_shifts += shift
		# self._encode_at_most_x_workshifts_per_y_days_binomial(6, 7)
		if not eval_window_lower_bound(nurse_shifts, 'O', 7, 7 - 6):
			raise RuntimeError(
				f"nurse id {nurse_id} failed at self._encode_at_most_x_workshifts_per_y_days_binomial(6, 7)")
		# self._encode_at_least_x_offdays_per_y_days_staircase(4, 18)
		if not eval_window_lower_bound(nurse_shifts, 'O', 14, 4):
			raise RuntimeError(
				f"nurse id {nurse_id} failed at self._encode_at_least_x_offdays_per_y_days(4, 18)")
		# self._encode_between_x_and_y_s_shifts_per_z_days(4, 8, ShiftEnum.EVENING_SHIFT, 14)
		if not eval_window(nurse_shifts, 'E', 14, 4, 8):
			raise RuntimeError(
				f"nurse id {nurse_id} failed at self._encode_between_x_and_y_s_shifts_per_z_days(4, 8, ShiftEnum.EVENING_SHIFT, 14)")
		# self._encode_between_x_and_y_workshifts_per_z_days(16, 18, 28)
		if not eval_window_upper_bound(nurse_shifts, 'O', 28, 28 - 20):
			raise RuntimeError(
				f"nurse id {nurse_id} failed at self._encode_between_x_and_y_workshifts_per_z_days(16, 18, 28)")
		# self._encode_at_most_x_s_shifts_per_y_days_binomial(2, ShiftEnum.NIGHT_SHIFT, 7)
		if not eval_window_upper_bound(nurse_shifts, 'N', 7, 2):
			raise RuntimeError(
				f"nurse id {nurse_id} failed at self._encode_at_most_x_s_shifts_per_y_days_binomial(2, ShiftEnum.NIGHT_SHIFT, 7)")
		# self._encode_at_least_x_s_shifts_per_y_days_binomial(1, ShiftEnum.NIGHT_SHIFT, 7)
		if not eval_window_lower_bound(nurse_shifts, 'N', 14, 1):
			raise RuntimeError(
				f"nurse id {nurse_id} failed at self._encode_at_least_x_s_shifts_per_y_days_binomial(1, ShiftEnum.NIGHT_SHIFT, 7)")
		# self._encode_between_x_and_y_s_shifts_per_z_days(2, 4, ShiftEnum.EVENING_SHIFT, 7)
		if not eval_window(nurse_shifts, 'E', 7, 2, 4):
			raise RuntimeError(
				f"nurse id {nurse_id} failed at self._encode_between_x_and_y_s_shifts_per_z_days(2, 4, ShiftEnum.EVENING_SHIFT, 7)")
		# self._encode_at_most_x_s_shifts_per_y_days_binomial(1, ShiftEnum.NIGHT_SHIFT, 2)
		if not eval_window_upper_bound(nurse_shifts, 'N', 2, 1):
			raise RuntimeError(
				f"nurse id {nurse_id} failed at self._encode_at_most_x_s_shifts_per_y_days_binomial(1, ShiftEnum.NIGHT_SHIFT, 2)")
	print("ok")


def main():
	to_test: list[str] = ["staircase"]
	# to_test: list[str] = ["staircase"]
	time_now = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
	nks = get_all_number_in_file("input_nurse_rostering.txt")
	excel_file_name = f"results_nurse_rostering_{time_now}.xlsx"
	writer: DataToXlsx = DataToXlsx(excel_file_name, to_test)
	timeout = 600
	for nk in nks:
		nurse = nk[0]
		day = nk[1] * 7
		writer.write_more('nurse', {'nurse': nurse})
		writer.write_more('day', {'day': day})
		for name in to_test:
			result_dict: dict[str, str | int | float] = {
				DataToXlsx.total_variable: "",
				DataToXlsx.clause: "",
				DataToXlsx.time: "",
				DataToXlsx.sat_status: ""
			}
			elapsed_time_ms, solver_return, num_var, num_clause = run_nurse_rostering(name, nurse, day, timeout)
			if elapsed_time_ms is None:
				result_dict[DataToXlsx.time] = "timeout"
			else:
				result_dict[DataToXlsx.time] = f"{elapsed_time_ms:.3f}"
			if num_var == -1:
				result_dict[DataToXlsx.total_variable] = ""
				result_dict[DataToXlsx.clause] = ""
			else:
				result_dict[DataToXlsx.total_variable] = num_var
				result_dict[DataToXlsx.clause] = num_clause
				result_dict[DataToXlsx.sat_status] = solver_return

			writer.write_more(name, result_dict)
		print()
	pass


if __name__ == '__main__':
	main()
	pass
