import os
import signal
import sys
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from datetime import datetime
from typing import Any

from src.encoding.staircase_encoding import StaircaseEncoding
from src.include.common import myrange_inclusive, cl


def handler(signum, frame):
	raise TimeoutError("Function execution time exceeded the limit")


def run(command: str) -> int:
	print(command)
	return os.system(command)


def eval_window(x: str, window_size: int, cap: int):
	now = 0
	for i in range(0, len(x)):
		if i >= window_size:
			if now > cap:
				return False
			now -= 1 if x[i - window_size] == '1' else 0
		now += 1 if x[i] == '1' else 0
	if now > cap:
		return False
	return True


def test(name: str, n: int, window_size: int, cap: int, time_limit: int) -> tuple[float | None, int, int]:
	signal.signal(signal.SIGALRM, handler)
	signal.alarm(time_limit)
	cannon_name = f"{name}_{n}_{window_size}_{cap}.cnf"
	# if not os.path.exists("/tmp/amk"):
	# 	os.makedirs("/tmp/amk")
	filename = f"/tmp/amk/{cannon_name}"
	print(f"{name} {n} {window_size} {cap}")
	added_var: int = -1
	num_clause: int = -1
	try:
		start_time = time.perf_counter()
		formula = []
		var = [_ for _ in myrange_inclusive(1, n)]

		if name == "staircase":
			added_var = StaircaseEncoding().staircase(var, window_size, cap, n, formula)
		elif name == "staircase_brute_nsc":
			added_var = StaircaseEncoding().staircase_brute_nsc(var, window_size, cap, n, formula)
		elif name == "staircase_brute_pblib":
			added_var = StaircaseEncoding().staircase_brute_pblib(var, window_size, cap, n, formula)
		num_clause = len(formula)
		with open(filename, "w") as file:
			print(f"p cnf {n + added_var} {len(formula)}", file=file)
			for line in formula:
				cl(line, file=file)
		# if "range" in name:
		# 	func = function_template.name_to_func_range(name)
		# 	num_var, num_clause = func(n, k, m, filename)
		# else:
		# 	func = function_template.name_to_func(name)
		# 	num_var, num_clause = func(n, k, filename)

		os.system(f"head -n1 {filename}")
		if not os.path.exists("/tmp/amk/kissat_output"):
			os.makedirs("/tmp/amk/kissat_output")
		run(f"kissat -n --time={time_limit} {filename} > /tmp/amk/kissat_output/output_{cannon_name}.txt")
		end_time = time.perf_counter()
		elapsed_time_ms = (end_time - start_time) * 1000
		print(f"took {elapsed_time_ms:.2f} (ms)")
	except TimeoutError as te:
		print(te)
		return None, added_var, num_clause
	except RecursionError as re:
		print(re)
		return None, added_var, num_clause
	except OSError as e:
		print(e)
		return None, added_var, num_clause
	finally:
		signal.alarm(0)
	print("-" * 100)
	return elapsed_time_ms, added_var + n, num_clause


def get_all_number_in_file(file: str) -> list[list[int]]:
	args = []
	file = open(f'{file}', 'r')
	for line in file:
		arg = [int(num) for num in line.split()]
		args.append(arg)
	file.close()
	return args


def number_to_column_letter(n: int) -> str:
	result = ""
	while n > 0:
		n, remainder = divmod(n - 1, 26)
		result = chr(65 + remainder) + result

	return result


def pos_2d_to_pos_excel(x: int, y: int) -> str:
	ret = number_to_column_letter(y) + str(x)
	# print(f"({x}, {y}) -> {ret}")
	return ret


def write_to_cell(cell, value: Any):
	cell.value = value
	cell.alignment = Alignment(horizontal="center")


class DataToXlsx:
	total_variable = 'total_variable'
	new_variable = 'new_variable'
	clause = 'clause'
	time = 'time'

	def __init__(self, excel_file_name: str, name_list: list[str]):
		if not os.path.exists("out/excel"):
			os.makedirs("out/excel")
		self.excel_file_name = "out/excel/" + excel_file_name
		if os.path.exists(self.excel_file_name):
			print(f"file {self.excel_file_name} exist! aborting...")
			sys.exit(2)
		self.book = Workbook()
		for sheet_name in self.book.sheetnames:
			self.book.remove(self.book[sheet_name])
		self.book.create_sheet('Results')
		self.sheet = self.book['Results']
		self.name_to_column: dict[str, int] = {
			'n': 1,
			'window_size': 2,
			'cap': 3,
		}
		self.row_dict: dict[str, int] = {
			'n': 3,
			'window_size': 3,
			'cap': 3
		}
		self.offset_dict: dict[str, int] = {
			'n': 0,
			'window_size': 0,
			'cap': 0,
			DataToXlsx.total_variable: 0,
			DataToXlsx.new_variable: 1,
			DataToXlsx.clause: 2,
			DataToXlsx.time: 3,
		}
		write_to_cell(self.sheet.cell(2, 1), 'n')
		write_to_cell(self.sheet.cell(2, 2), 'window_size')
		write_to_cell(self.sheet.cell(2, 3), 'cap')

		for i in range(len(name_list)):
			y: int = 4 + i * 4
			self.name_to_column[name_list[i]] = y
			self.sheet.merge_cells(f"{pos_2d_to_pos_excel(1, y)}:{pos_2d_to_pos_excel(1, y + 3)}")
			write_to_cell(self.sheet.cell(1, y), name_list[i])
			for name_offset, offset in self.offset_dict.items():
				write_to_cell(self.sheet.cell(2, y + offset), name_offset)
			self.row_dict[name_list[i]] = 3
		self.book.save(self.excel_file_name)

	def get_column(self, name: str) -> int:
		if not (name in self.name_to_column):
			raise RuntimeError(f"{name} not in name_to_column")
		return self.name_to_column.get(name)

	def get_row(self, name: str) -> int:
		ret: int = self.row_dict.get(name)
		self.row_dict[name] += 1
		return ret

	def write_more(self, name: str, result: dict[str, str | int | float]) -> None:
		row: int = self.get_row(name)
		col: int = self.get_column(name)
		for name_offset, value in result.items():
			offset: int = self.offset_dict[name_offset]
			write_to_cell(self.sheet.cell(row, col + offset), value)
		self.book.save(self.excel_file_name)


def main():
	to_test: list[str] = ["staircase"]
	time_now = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
	nks = get_all_number_in_file("input.txt")
	excel_file_name = f"results_{time_now}.xlsx"
	writer: DataToXlsx = DataToXlsx(excel_file_name, to_test)
	timeout = 600
	for nk in nks:
		n = nk[0]
		window_size = nk[1]
		cap = nk[2]
		writer.write_more('n', {'n': n})
		writer.write_more('window_size', {'window_size': window_size})
		writer.write_more('cap', {'cap': cap})
		for name in to_test:
			result_dict: dict[str, str | int | float] = {
				DataToXlsx.total_variable: "",
				DataToXlsx.new_variable: "",
				DataToXlsx.clause: "",
				DataToXlsx.time: "",
			}
			elapsed_time_ms, num_var, num_clause = test(name, n, window_size, cap, timeout)
			if elapsed_time_ms is None:
				result_dict[DataToXlsx.time] = "timeout"
			else:
				result_dict[DataToXlsx.time] = f"{elapsed_time_ms:.3f}"
			if num_var == -1:
				result_dict[DataToXlsx.total_variable] = ""
				result_dict[DataToXlsx.new_variable] = ""
				result_dict[DataToXlsx.clause] = ""
			else:
				result_dict[DataToXlsx.total_variable] = num_var
				result_dict[DataToXlsx.new_variable] = num_var - n
				result_dict[DataToXlsx.clause] = num_clause

			writer.write_more(name, result_dict)
		print()
	pass


if __name__ == '__main__':
	main()
